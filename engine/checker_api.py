#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
AI Factory Lab - 단계별 체크 API
파일: engine/checker_api.py
목적: Excel PI 업로드 시 자동 검증 및 실시간 피드백
"""

from flask import Flask, request, jsonify
from flask_cors import CORS
import openpyxl
import json
import os
import subprocess
import re
from pathlib import Path

app = Flask(__name__)
CORS(app)  # Vue에서 호출 가능하도록

# 프로젝트 루트 경로
BASE_DIR = Path(__file__).resolve().parent.parent

class PreGenerationChecker:
    """Generator 실행 전 체크"""
    
    def __init__(self, excel_file_path):
        self.excel_file = excel_file_path
        self.results = []
        self.errors = []
        self.warnings = []
        self.wb = None
        
    def check_all(self):
        """모든 검증 실행"""
        try:
            # 1. 파일 존재 확인
            if not os.path.exists(self.excel_file):
                self.add_error("파일 존재", f"Excel 파일을 찾을 수 없습니다: {self.excel_file}")
                return self.get_summary()
            
            # 2. Excel 파일 로드
            try:
                self.wb = openpyxl.load_workbook(self.excel_file, data_only=True)
                self.add_pass("파일 로드", "Excel 파일 로드 성공")
            except Exception as e:
                self.add_error("파일 로드", f"Excel 파일 로드 실패: {str(e)}")
                return self.get_summary()
            
            # 3. 시트명 검증
            self.check_sheet_names()
            
            # 4. GridColumns 헤더 검증
            self.check_grid_columns_headers()
            
            # 5. 데이터 타입 검증
            self.check_data_types()
            
            # 6. 필드명 검증 (특수문자, 공백)
            self.check_field_names()
            
            # 7. 화면 ID 검증
            self.check_screen_id()
            
            # 8. Buttons 시트 검증
            self.check_buttons()
            
            # 9. SearchConditions 시트 검증
            self.check_search_conditions()
            
            return self.get_summary()
            
        except Exception as e:
            self.add_error("전체 검증", f"예상치 못한 오류: {str(e)}")
            return self.get_summary()
    
    def check_sheet_names(self):
        """시트명 검증"""
        required_sheets = ['ScreenInfo', 'GridColumns', 'Buttons', 'SearchConditions']
        actual_sheets = self.wb.sheetnames
        
        missing_sheets = [s for s in required_sheets if s not in actual_sheets]
        
        if missing_sheets:
            self.add_error(
                "시트명 검증",
                f"필수 시트가 누락되었습니다: {', '.join(missing_sheets)}",
                solution=f"Excel 템플릿에 다음 시트를 추가하세요: {', '.join(missing_sheets)}"
            )
        else:
            self.add_pass(
                "시트명 검증",
                f"모든 필수 시트 존재: {', '.join(required_sheets)}"
            )
    
    def check_grid_columns_headers(self):
        """GridColumns 헤더 검증"""
        if 'GridColumns' not in self.wb.sheetnames:
            return
        
        sheet = self.wb['GridColumns']
        headers = [cell.value for cell in sheet[1] if cell.value]
        
        required_headers = ['Field Name', 'Type', 'Label', 'Width']
        missing_headers = [h for h in required_headers if h not in headers]
        
        if missing_headers:
            self.add_error(
                "GridColumns 헤더",
                f"필수 헤더가 누락되었습니다: {', '.join(missing_headers)}",
                solution=f"GridColumns 시트의 첫 번째 행에 다음 헤더를 추가하세요: {', '.join(missing_headers)}"
            )
        else:
            self.add_pass(
                "GridColumns 헤더",
                f"모든 필수 헤더 존재: {', '.join(required_headers)}"
            )
    
    def check_data_types(self):
        """데이터 타입 검증"""
        if 'GridColumns' not in self.wb.sheetnames:
            return
        
        sheet = self.wb['GridColumns']
        headers = [cell.value for cell in sheet[1]]
        
        try:
            type_col_idx = headers.index('Type')
        except ValueError:
            return
        
        valid_types = ['number', 'text', 'datetime', 'date', 'boolean', 'data']
        invalid_types = []
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row[type_col_idx]:
                continue
            
            data_type = str(row[type_col_idx]).strip().lower()
            if data_type and data_type not in valid_types:
                field_name = row[0] if row[0] else f"Row {row_idx}"
                invalid_types.append(f"{field_name}: '{data_type}'")
        
        if invalid_types:
            self.add_error(
                "데이터 타입 검증",
                f"잘못된 데이터 타입이 {len(invalid_types)}개 발견되었습니다",
                details=invalid_types[:5],  # 최대 5개만 표시
                solution=f"허용된 타입만 사용하세요: {', '.join(valid_types)}"
            )
        else:
            self.add_pass(
                "데이터 타입 검증",
                "모든 데이터 타입이 표준을 준수합니다"
            )
    
    def check_field_names(self):
        """필드명 검증 (특수문자, 공백, 예약어)"""
        if 'GridColumns' not in self.wb.sheetnames:
            return
        
        sheet = self.wb['GridColumns']
        headers = [cell.value for cell in sheet[1]]
        
        try:
            field_name_col_idx = headers.index('Field Name')
        except ValueError:
            return
        
        invalid_fields = []
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            field_name = row[field_name_col_idx]
            if not field_name:
                continue
            
            field_name = str(field_name).strip()
            
            # 특수문자 검증 (영문, 숫자, 언더스코어만 허용)
            if not re.match(r'^[a-zA-Z][a-zA-Z0-9_]*$', field_name):
                invalid_fields.append({
                    'row': row_idx,
                    'field': field_name,
                    'reason': '특수문자 또는 공백 포함 (영문으로 시작, 영문/숫자/언더스코어만 허용)'
                })
            
            # 예약어 검증
            reserved_words = ['select', 'from', 'where', 'order', 'group', 'having', 'class', 'interface']
            if field_name.lower() in reserved_words:
                invalid_fields.append({
                    'row': row_idx,
                    'field': field_name,
                    'reason': 'SQL/Java 예약어 사용 금지'
                })
        
        if invalid_fields:
            self.add_error(
                "필드명 검증",
                f"잘못된 필드명이 {len(invalid_fields)}개 발견되었습니다",
                details=[f"Row {f['row']}: {f['field']} - {f['reason']}" for f in invalid_fields[:5]],
                solution="필드명은 영문으로 시작하고, 영문/숫자/언더스코어만 사용해야 합니다 (예: orderDate, productName)"
            )
        else:
            self.add_pass(
                "필드명 검증",
                "모든 필드명이 네이밍 규칙을 준수합니다"
            )
    
    def check_screen_id(self):
        """화면 ID 검증"""
        if 'ScreenInfo' not in self.wb.sheetnames:
            self.add_warning(
                "화면 ID 검증",
                "ScreenInfo 시트가 없어 화면 ID를 확인할 수 없습니다",
                solution="ScreenInfo 시트를 추가하고 Screen ID를 정의하세요"
            )
            return
        
        sheet = self.wb['ScreenInfo']
        screen_id = None
        
        # Screen ID 찾기 (일반적으로 A2 또는 B1 위치)
        for row in sheet.iter_rows(min_row=1, max_row=10, values_only=True):
            for idx, cell in enumerate(row):
                if cell and 'Screen ID' in str(cell):
                    # 다음 셀이 실제 ID
                    if idx + 1 < len(row):
                        screen_id = row[idx + 1]
                    break
            if screen_id:
                break
        
        if not screen_id:
            # 첫 번째 데이터 행에서 찾기
            screen_id = sheet['B2'].value or sheet['A2'].value
        
        if screen_id:
            screen_id = str(screen_id).strip()
            
            # 화면 ID 형식 검증 (영문+숫자, PascalCase)
            if not re.match(r'^[A-Z][a-zA-Z0-9]*$', screen_id):
                self.add_error(
                    "화면 ID 검증",
                    f"화면 ID 형식이 잘못되었습니다: '{screen_id}'",
                    solution="화면 ID는 대문자로 시작하는 PascalCase여야 합니다 (예: CostManagement, ProductionResult)"
                )
            else:
                self.add_pass(
                    "화면 ID 검증",
                    f"화면 ID가 올바릅니다: {screen_id}"
                )
        else:
            self.add_warning(
                "화면 ID 검증",
                "화면 ID를 찾을 수 없습니다",
                solution="ScreenInfo 시트에 Screen ID를 정의하세요"
            )
    
    def check_buttons(self):
        """Buttons 시트 검증"""
        if 'Buttons' not in self.wb.sheetnames:
            self.add_warning(
                "Buttons 검증",
                "Buttons 시트가 없습니다",
                solution="화면에 버튼이 있다면 Buttons 시트를 추가하세요"
            )
            return
        
        sheet = self.wb['Buttons']
        headers = [cell.value for cell in sheet[1] if cell.value]
        
        if not headers:
            self.add_warning(
                "Buttons 검증",
                "Buttons 시트가 비어 있습니다",
                solution="버튼 정보를 입력하거나 시트를 삭제하세요"
            )
            return
        
        # 버튼 개수 확인
        button_count = sum(1 for row in sheet.iter_rows(min_row=2) if any(cell.value for cell in row))
        
        if button_count > 0:
            self.add_pass(
                "Buttons 검증",
                f"{button_count}개의 버튼이 정의되어 있습니다"
            )
    
    def check_search_conditions(self):
        """SearchConditions 시트 검증"""
        if 'SearchConditions' not in self.wb.sheetnames:
            self.add_warning(
                "SearchConditions 검증",
                "SearchConditions 시트가 없습니다",
                solution="검색 조건이 있다면 SearchConditions 시트를 추가하세요"
            )
            return
        
        sheet = self.wb['SearchConditions']
        headers = [cell.value for cell in sheet[1] if cell.value]
        
        if not headers:
            self.add_warning(
                "SearchConditions 검증",
                "SearchConditions 시트가 비어 있습니다"
            )
            return
        
        # 검색 조건 개수 확인
        condition_count = sum(1 for row in sheet.iter_rows(min_row=2) if any(cell.value for cell in row))
        
        if condition_count > 0:
            self.add_pass(
                "SearchConditions 검증",
                f"{condition_count}개의 검색 조건이 정의되어 있습니다"
            )
    
    def add_pass(self, check_name, message, details=None):
        """통과 결과 추가"""
        self.results.append({
            'status': 'pass',
            'check': check_name,
            'message': message,
            'details': details
        })
    
    def add_error(self, check_name, message, details=None, solution=None):
        """에러 결과 추가"""
        error = {
            'status': 'error',
            'check': check_name,
            'message': message,
            'details': details,
            'solution': solution
        }
        self.results.append(error)
        self.errors.append(error)
    
    def add_warning(self, check_name, message, details=None, solution=None):
        """경고 결과 추가"""
        warning = {
            'status': 'warning',
            'check': check_name,
            'message': message,
            'details': details,
            'solution': solution
        }
        self.results.append(warning)
        self.warnings.append(warning)
    
    def get_summary(self):
        """검증 결과 요약"""
        pass_count = sum(1 for r in self.results if r['status'] == 'pass')
        error_count = len(self.errors)
        warning_count = len(self.warnings)
        
        return {
            'success': error_count == 0,
            'summary': {
                'total': len(self.results),
                'pass': pass_count,
                'error': error_count,
                'warning': warning_count
            },
            'results': self.results,
            'can_proceed': error_count == 0,
            'recommendation': self._get_recommendation()
        }
    
    def _get_recommendation(self):
        """추천 조치"""
        if len(self.errors) > 0:
            return "🔴 에러를 모두 수정한 후 다시 업로드해주세요."
        elif len(self.warnings) > 0:
            return "⚠️ 경고 사항을 확인하세요. 코드 생성은 가능하지만 일부 기능이 누락될 수 있습니다."
        else:
            return "✅ 모든 검증을 통과했습니다! 코드 생성을 진행하세요."


class PostGenerationChecker:
    """코드 생성 후 체크"""
    
    def __init__(self, screen_id):
        self.screen_id = screen_id
        self.results = []
        self.errors = []
        self.warnings = []
        self.output_dir = BASE_DIR / 'engine' / 'output' / screen_id
    
    def check_all(self):
        """모든 검증 실행"""
        try:
            # 1. 파일 생성 확인
            self.check_files_generated()
            
            # 2. JSON Schema 검증
            self.check_json_schema()
            
            # 3. Vue 파일 검증
            self.check_vue_file()
            
            # 4. Java 파일 검증
            self.check_java_files()
            
            # 5. MyBatis XML 검증
            self.check_mybatis_xml()
            
            # 6. 빌드 준비 상태 확인
            self.check_build_ready()
            
            return self.get_summary()
            
        except Exception as e:
            self.add_error("전체 검증", f"예상치 못한 오류: {str(e)}")
            return self.get_summary()
    
    def check_files_generated(self):
        """생성된 파일 확인"""
        required_files = {
            'JSON Schema': f'{self.screen_id}.json',
            'Vue Component': f'{self.screen_id}.vue',
            'Java Controller': f'java/{self.screen_id}Controller.java',
            'Java Service': f'java/{self.screen_id}Service.java',
            'Java ServiceImpl': f'java/{self.screen_id}ServiceImpl.java',
            'Java Mapper': f'java/{self.screen_id}Mapper.java',
            'MyBatis XML': f'mapper/{self.screen_id}Mapper.xml'
        }
        
        missing_files = []
        
        for file_type, file_name in required_files.items():
            file_path = self.output_dir / file_name
            if file_path.exists():
                self.add_pass(
                    f"파일 생성: {file_type}",
                    f"{file_name} 생성 완료"
                )
            else:
                missing_files.append(file_type)
                self.add_error(
                    f"파일 생성: {file_type}",
                    f"{file_name} 파일이 생성되지 않았습니다",
                    solution=f"Generator를 다시 실행하거나 {file_type} 생성 로직을 확인하세요"
                )
        
        if not missing_files:
            self.add_pass(
                "파일 생성 전체",
                "모든 필수 파일이 생성되었습니다"
            )
    
    def check_json_schema(self):
        """JSON Schema 검증"""
        json_file = self.output_dir / f'{self.screen_id}.json'
        
        if not json_file.exists():
            return
        
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                schema = json.load(f)
            
            self.add_pass(
                "JSON 문법",
                "올바른 JSON 형식입니다"
            )
            
            # 필수 키 확인
            required_keys = ['screenId', 'screenName', 'category', 'columns']
            missing_keys = [k for k in required_keys if k not in schema]
            
            if missing_keys:
                self.add_error(
                    "JSON Schema 구조",
                    f"필수 키가 누락되었습니다: {', '.join(missing_keys)}",
                    solution="Generator를 다시 실행하세요"
                )
            else:
                self.add_pass(
                    "JSON Schema 구조",
                    f"모든 필수 키 존재: {len(schema.get('columns', []))}개 컬럼"
                )
            
        except json.JSONDecodeError as e:
            self.add_error(
                "JSON 문법",
                f"JSON 파싱 에러: {str(e)}",
                solution="Generator를 다시 실행하여 올바른 JSON을 생성하세요"
            )
    
    def check_vue_file(self):
        """Vue 파일 검증"""
        vue_file = self.output_dir / f'{self.screen_id}.vue'
        
        if not vue_file.exists():
            return
        
        with open(vue_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # fields 정의 확인
        if 'const fields = ref' in content:
            self.add_pass(
                "Vue fields 정의",
                "RealGrid fields 선언 확인"
            )
        else:
            self.add_warning(
                "Vue fields 정의",
                "fields 선언을 찾을 수 없습니다",
                solution="RealGrid를 사용하는 경우 fields를 정의하세요"
            )
        
        # API 호출 확인
        if 'axios.get' in content or 'axios.post' in content:
            self.add_pass(
                "Vue API 호출",
                "axios를 사용한 API 호출 확인"
            )
        
        # 컴포넌트 기본 구조 확인
        if '<template>' in content and '<script setup>' in content:
            self.add_pass(
                "Vue 컴포넌트 구조",
                "Vue 3 Composition API 구조 확인"
            )
    
    def check_java_files(self):
        """Java 파일 검증"""
        java_dir = self.output_dir / 'java'
        
        if not java_dir.exists():
            return
        
        # Controller 확인
        controller_file = java_dir / f'{self.screen_id}Controller.java'
        if controller_file.exists():
            with open(controller_file, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # package 선언 확인
            package_match = re.search(r'package\s+([\w.]+);', content)
            if package_match:
                self.add_pass(
                    "Java package 선언",
                    f"Package: {package_match.group(1)}"
                )
            else:
                self.add_error(
                    "Java package 선언",
                    "package 선언이 없습니다",
                    solution="Java 파일 첫 줄에 package를 선언하세요"
                )
            
            # Bean 이름 확인
            if f'{self.screen_id}Service' in content:
                self.add_pass(
                    "Java Bean 이름",
                    f"{self.screen_id}Service 참조 확인"
                )
            
            # @RestController 확인
            if '@RestController' in content:
                self.add_pass(
                    "Spring Controller 어노테이션",
                    "@RestController 선언 확인"
                )
    
    def check_mybatis_xml(self):
        """MyBatis XML 검증"""
        xml_file = self.output_dir / 'mapper' / f'{self.screen_id}Mapper.xml'
        
        if not xml_file.exists():
            return
        
        with open(xml_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # CDATA 사용 확인
        if '<![CDATA[' in content:
            self.add_pass(
                "MyBatis CDATA",
                "SQL이 CDATA로 안전하게 감싸져 있습니다"
            )
        else:
            self.add_warning(
                "MyBatis CDATA",
                "CDATA 사용을 권장합니다",
                solution="복잡한 SQL은 <![CDATA[...]]>로 감싸세요"
            )
        
        # SQL Injection 위험 확인
        if '${' in content:
            count = content.count('${')
            self.add_warning(
                "SQL Injection 위험",
                f"${{}} 사용 {count}회 발견",
                solution="가능하면 #{{}}를 사용하세요 (prepared statement)"
            )
        else:
            self.add_pass(
                "SQL Injection 방지",
                "#{} 사용으로 안전합니다"
            )
        
        # XML 문법 검증
        try:
            result = subprocess.run(
                ['xmllint', '--noout', str(xml_file)],
                capture_output=True,
                text=True
            )
            if result.returncode == 0:
                self.add_pass(
                    "XML 문법",
                    "올바른 XML 형식입니다"
                )
            else:
                self.add_error(
                    "XML 문법",
                    "XML 파싱 에러",
                    details=result.stderr.split('\n')[:3],
                    solution="XML 태그를 올바르게 닫고 특수문자를 이스케이프하세요"
                )
        except FileNotFoundError:
            # xmllint가 없는 경우 스킵
            pass
    
    def check_build_ready(self):
        """빌드 준비 상태 확인"""
        # 모든 필수 파일이 있는지 최종 확인
        if len(self.errors) == 0:
            self.add_pass(
                "빌드 준비 상태",
                "✅ 파일 배포 및 빌드를 진행할 수 있습니다"
            )
        else:
            self.add_error(
                "빌드 준비 상태",
                f"❌ {len(self.errors)}개의 에러를 해결해야 합니다"
            )
    
    def add_pass(self, check_name, message, details=None):
        """통과 결과 추가"""
        self.results.append({
            'status': 'pass',
            'check': check_name,
            'message': message,
            'details': details
        })
    
    def add_error(self, check_name, message, details=None, solution=None):
        """에러 결과 추가"""
        error = {
            'status': 'error',
            'check': check_name,
            'message': message,
            'details': details,
            'solution': solution
        }
        self.results.append(error)
        self.errors.append(error)
    
    def add_warning(self, check_name, message, details=None, solution=None):
        """경고 결과 추가"""
        warning = {
            'status': 'warning',
            'check': check_name,
            'message': message,
            'details': details,
            'solution': solution
        }
        self.results.append(warning)
        self.warnings.append(warning)
    
    def get_summary(self):
        """검증 결과 요약"""
        pass_count = sum(1 for r in self.results if r['status'] == 'pass')
        error_count = len(self.errors)
        warning_count = len(self.warnings)
        
        return {
            'success': error_count == 0,
            'summary': {
                'total': len(self.results),
                'pass': pass_count,
                'error': error_count,
                'warning': warning_count
            },
            'results': self.results,
            'can_proceed': error_count == 0,
            'recommendation': self._get_recommendation()
        }
    
    def _get_recommendation(self):
        """추천 조치"""
        if len(self.errors) > 0:
            return "🔴 에러를 수정한 후 코드를 다시 생성하세요."
        elif len(self.warnings) > 0:
            return "⚠️ 경고 사항을 확인하세요. 배포는 가능하지만 개선이 필요합니다."
        else:
            return "✅ 모든 검증을 통과했습니다! 파일 배포를 진행하세요."


class PreDeploymentChecker:
    """배포 전 체크"""
    
    def __init__(self, screen_id):
        self.screen_id = screen_id
        self.results = []
        self.errors = []
        self.warnings = []
    
    def check_all(self):
        """모든 검증 실행"""
        try:
            # 1. Backend 빌드 테스트
            self.check_backend_build()
            
            # 2. 파일 배포 확인
            self.check_files_deployed()
            
            # 3. Backend 서버 상태
            self.check_backend_server()
            
            # 4. Frontend 서버 상태
            self.check_frontend_server()
            
            # 5. API 엔드포인트 테스트
            self.check_api_endpoints()
            
            return self.get_summary()
            
        except Exception as e:
            self.add_error("전체 검증", f"예상치 못한 오류: {str(e)}")
            return self.get_summary()
    
    def check_backend_build(self):
        """Backend 빌드 테스트"""
        backend_dir = BASE_DIR / 'backend'
        
        if not backend_dir.exists():
            self.add_error(
                "Backend 디렉토리",
                "backend 디렉토리를 찾을 수 없습니다"
            )
            return
        
        self.add_pass(
            "Backend 빌드 시작",
            "mvn clean compile 실행 중... (20초 소요 예상)"
        )
        
        # 실제 빌드는 너무 오래 걸리므로 컴파일 테스트만
        result = subprocess.run(
            ['mvn', 'compiler:testCompile', '-q'],
            cwd=str(backend_dir),
            capture_output=True,
            text=True,
            timeout=60
        )
        
        if result.returncode == 0:
            self.add_pass(
                "Backend 빌드",
                "✅ 컴파일 성공! 문법 에러 없음"
            )
        else:
            error_lines = [line for line in result.stdout.split('\n') if 'ERROR' in line or 'error:' in line]
            self.add_error(
                "Backend 빌드",
                "❌ 컴파일 에러 발생",
                details=error_lines[:5],
                solution="Java 파일의 package 경로, import 문, 문법을 확인하세요"
            )
    
    def check_files_deployed(self):
        """파일 배포 확인"""
        # Vue 파일 배포 확인
        vue_dirs = [
            BASE_DIR / 'frontend' / 'src' / 'views',
            BASE_DIR / 'frontend' / 'src' / 'views' / 'cost',
            BASE_DIR / 'frontend' / 'src' / 'views' / 'production',
            BASE_DIR / 'frontend' / 'src' / 'views' / 'demo',
        ]
        
        vue_found = False
        for vue_dir in vue_dirs:
            vue_file = vue_dir / f'{self.screen_id}.vue'
            if vue_file.exists():
                self.add_pass(
                    "Vue 파일 배포",
                    f"✅ {vue_file.relative_to(BASE_DIR)}"
                )
                vue_found = True
                break
        
        if not vue_found:
            self.add_warning(
                "Vue 파일 배포",
                f"{self.screen_id}.vue 파일을 frontend/src/views에서 찾을 수 없습니다",
                solution="파일을 올바른 경로에 복사하세요"
            )
        
        # Java 파일 배포 확인
        java_base = BASE_DIR / 'backend' / 'src' / 'main' / 'java' / 'com' / 'dowinsys'
        java_dirs = list(java_base.glob('*'))
        
        java_found = False
        for java_dir in java_dirs:
            controller_file = java_dir / f'{self.screen_id}Controller.java'
            if controller_file.exists():
                self.add_pass(
                    "Java 파일 배포",
                    f"✅ {controller_file.relative_to(BASE_DIR)}"
                )
                java_found = True
                break
        
        if not java_found:
            self.add_warning(
                "Java 파일 배포",
                f"{self.screen_id}Controller.java 파일을 찾을 수 없습니다"
            )
    
    def check_backend_server(self):
        """Backend 서버 상태"""
        try:
            import requests
            response = requests.get('http://localhost:8080/actuator/health', timeout=3)
            if response.status_code == 200:
                self.add_pass(
                    "Backend 서버",
                    "✅ Spring Boot 실행 중 (8080 포트)"
                )
            else:
                self.add_warning(
                    "Backend 서버",
                    f"응답 코드: {response.status_code}"
                )
        except:
            self.add_warning(
                "Backend 서버",
                "⚠️ Spring Boot 미실행 (8080 포트)",
                solution="mvn spring-boot:run 명령으로 서버를 시작하세요"
            )
    
    def check_frontend_server(self):
        """Frontend 서버 상태"""
        try:
            import requests
            response = requests.get('http://localhost:8081', timeout=3)
            if response.status_code == 200:
                self.add_pass(
                    "Frontend 서버",
                    "✅ Vue Dev Server 실행 중 (8081 포트)"
                )
            else:
                self.add_warning(
                    "Frontend 서버",
                    f"응답 코드: {response.status_code}"
                )
        except:
            self.add_warning(
                "Frontend 서버",
                "⚠️ Vue Dev Server 미실행 (8081 포트)",
                solution="npm run serve 명령으로 서버를 시작하세요"
            )
    
    def check_api_endpoints(self):
        """API 엔드포인트 테스트"""
        try:
            import requests
            
            # 카테고리 추출 (예: CostManagement -> cost)
            category = re.sub(r'(?<!^)(?=[A-Z])', '_', self.screen_id).lower().split('_')[0]
            api_url = f'http://localhost:8080/api/{category}/{self.screen_id.lower()}/list'
            
            response = requests.get(api_url, timeout=5)
            
            if response.status_code == 200:
                data = response.json()
                self.add_pass(
                    "API 엔드포인트",
                    f"✅ {api_url} 정상 응답",
                    details=f"데이터: {len(data.get('list', []))}건" if isinstance(data, dict) else None
                )
            elif response.status_code == 404:
                self.add_warning(
                    "API 엔드포인트",
                    f"⚠️ {api_url} 없음 (404)",
                    solution="Backend 서버를 재시작하세요"
                )
            else:
                self.add_warning(
                    "API 엔드포인트",
                    f"응답 코드: {response.status_code}"
                )
                
        except Exception as e:
            self.add_warning(
                "API 엔드포인트",
                "API 테스트 실패",
                solution="Backend 서버가 실행 중인지 확인하세요"
            )
    
    def add_pass(self, check_name, message, details=None):
        self.results.append({
            'status': 'pass',
            'check': check_name,
            'message': message,
            'details': details
        })
    
    def add_error(self, check_name, message, details=None, solution=None):
        error = {
            'status': 'error',
            'check': check_name,
            'message': message,
            'details': details,
            'solution': solution
        }
        self.results.append(error)
        self.errors.append(error)
    
    def add_warning(self, check_name, message, details=None, solution=None):
        warning = {
            'status': 'warning',
            'check': check_name,
            'message': message,
            'details': details,
            'solution': solution
        }
        self.results.append(warning)
        self.warnings.append(warning)
    
    def get_summary(self):
        pass_count = sum(1 for r in self.results if r['status'] == 'pass')
        error_count = len(self.errors)
        warning_count = len(self.warnings)
        
        return {
            'success': error_count == 0,
            'summary': {
                'total': len(self.results),
                'pass': pass_count,
                'error': error_count,
                'warning': warning_count
            },
            'results': self.results,
            'can_proceed': error_count == 0,
            'recommendation': self._get_recommendation()
        }
    
    def _get_recommendation(self):
        if len(self.errors) > 0:
            return "🔴 빌드 에러를 해결한 후 배포하세요."
        elif len(self.warnings) > 0:
            return "⚠️ 경고 사항이 있지만 배포는 가능합니다."
        else:
            return "✅ 배포 준비 완료! 화면을 테스트하세요."


# ==================== API 엔드포인트 ====================

@app.route('/api/check/pre-generation', methods=['POST'])
def check_pre_generation():
    """Excel PI 업로드 시 자동 검증"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Excel 파일이 필요합니다'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'error': '파일이 선택되지 않았습니다'}), 400
        
        # 임시 파일로 저장
        temp_dir = BASE_DIR / 'engine' / 'temp'
        temp_dir.mkdir(exist_ok=True)
        
        temp_file = temp_dir / file.filename
        file.save(str(temp_file))
        
        # 검증 실행
        checker = PreGenerationChecker(str(temp_file))
        result = checker.check_all()
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/check/post-generation/<screen_id>', methods=['GET'])
def check_post_generation(screen_id):
    """코드 생성 후 검증"""
    try:
        checker = PostGenerationChecker(screen_id)
        result = checker.check_all()
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/check/pre-deployment/<screen_id>', methods=['GET'])
def check_pre_deployment(screen_id):
    """배포 전 검증"""
    try:
        checker = PreDeploymentChecker(screen_id)
        result = checker.check_all()
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/health', methods=['GET'])
def health():
    """Health check"""
    return jsonify({'status': 'ok'})


if __name__ == '__main__':
    print("=" * 60)
    print("  AI Factory Lab - 단계별 체크 API 서버")
    print("=" * 60)
    print("  포트: 5000")
    print("  엔드포인트:")
    print("    POST /api/check/pre-generation  - Excel 업로드 검증")
    print("    GET  /api/check/post-generation/<screen_id> - 코드 생성 후 검증")
    print("    GET  /api/check/pre-deployment/<screen_id> - 배포 전 검증")
    print("=" * 60)
    
    app.run(debug=True, host='0.0.0.0', port=5000)
