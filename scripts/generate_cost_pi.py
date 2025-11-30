#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
원가 관리 화면 PI 생성 스크립트
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_cost_management_pi():
    """원가 관리 화면 Excel PI 생성"""
    
    wb = openpyxl.Workbook()
    
    # Sheet 1: 기본 정보
    ws1 = wb.active
    ws1.title = "01_BasicInfo"
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # 기본 정보 입력
    basic_info = [
        ["항목명", "값"],
        ["화면ID", "CostManagement"],
        ["화면명(한글)", "원가 관리"],
        ["화면명(영문)", "Cost Management"],
        ["카테고리", "cost"],
        ["설명", "품목별/공장별 원가 정보를 조회하고 관리하는 화면입니다. Excel 업로드/다운로드 기능 포함"],
        ["테이블명", "new_doi_cost_master"],
        ["Primary Key", "cost_id"],
        ["조회 기능", "Y"],
        ["행 추가 가능", "Y"],
        ["행 삭제 가능", "Y"],
        ["저장 기능", "Y"],
        ["Excel 업로드", "Y"],
        ["Excel 다운로드", "Y"],
        ["일괄 승인 기능", "Y"],
    ]
    
    for row in basic_info:
        ws1.append(row)
    
    # 헤더 스타일 적용
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Sheet 2: 그리드 컬럼
    ws2 = wb.create_sheet("02_GridColumns")
    
    grid_columns = [
        ["Field Name", "Header Text", "Data Type", "Width", "Editable", "Alignment", "Excel Mapping"],
        ["rowNum", "No", "number", "50", "N", "center", "N"],
        ["costId", "원가ID", "string", "100", "N", "left", "N"],
        ["factoryCd", "공장코드", "string", "80", "Y", "center", "Y"],
        ["factoryNm", "공장명", "string", "120", "N", "left", "Y"],
        ["itemCd", "품목코드", "string", "100", "Y", "center", "Y"],
        ["itemNm", "품목명", "string", "200", "N", "left", "Y"],
        ["itemSpec", "규격", "string", "150", "Y", "left", "Y"],
        ["unit", "단위", "string", "60", "Y", "center", "Y"],
        ["costType", "원가유형", "string", "100", "Y", "center", "Y"],
        ["baseYear", "기준년도", "string", "80", "Y", "center", "Y"],
        ["baseMonth", "기준월", "string", "70", "Y", "center", "Y"],
        ["materialCost", "재료비", "number", "100", "Y", "right", "Y"],
        ["laborCost", "노무비", "number", "100", "Y", "right", "Y"],
        ["overheadCost", "경비", "number", "100", "Y", "right", "Y"],
        ["totalCost", "총원가", "number", "120", "N", "right", "Y"],
        ["profitRate", "이익률(%)", "number", "90", "Y", "right", "Y"],
        ["sellingPrice", "판매가", "number", "120", "Y", "right", "Y"],
        ["currency", "통화", "string", "70", "Y", "center", "Y"],
        ["exchangeRate", "환율", "number", "100", "Y", "right", "Y"],
        ["supplierCd", "공급업체코드", "string", "100", "Y", "center", "Y"],
        ["supplierNm", "공급업체명", "string", "150", "N", "left", "Y"],
        ["validFrom", "유효시작일", "date", "100", "Y", "center", "Y"],
        ["validTo", "유효종료일", "date", "100", "Y", "center", "Y"],
        ["status", "상태", "string", "80", "Y", "center", "Y"],
        ["approvalStatus", "승인상태", "string", "80", "N", "center", "Y"],
        ["approvalDate", "승인일자", "datetime", "140", "N", "center", "Y"],
        ["approvalUser", "승인자", "string", "100", "N", "left", "Y"],
        ["remark", "비고", "string", "200", "Y", "left", "Y"],
        ["createDate", "생성일시", "datetime", "140", "N", "center", "N"],
        ["createUser", "생성자", "string", "100", "N", "left", "N"],
        ["updateDate", "수정일시", "datetime", "140", "N", "center", "N"],
        ["updateUser", "수정자", "string", "100", "N", "left", "N"],
    ]
    
    for row in grid_columns:
        ws2.append(row)
    
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Sheet 3: 검색 조건
    ws3 = wb.create_sheet("03_SearchConditions")
    
    search_conditions = [
        ["Field Name", "Label", "Type", "Required", "Default Value", "Options"],
        ["baseYearFrom", "기준년도(시작)", "text", "N", "", ""],
        ["baseYearTo", "기준년도(종료)", "text", "N", "", ""],
        ["baseMonth", "기준월", "select", "N", "", "01:1월,02:2월,03:3월,04:4월,05:5월,06:6월,07:7월,08:8월,09:9월,10:10월,11:11월,12:12월"],
        ["factoryCd", "공장", "select", "N", "", "F001:본사공장,F002:2공장,F003:3공장"],
        ["itemCd", "품목코드", "text", "N", "", ""],
        ["itemNm", "품목명", "text", "N", "", ""],
        ["costType", "원가유형", "select", "N", "", "STD:표준원가,ACT:실제원가,EST:견적원가"],
        ["supplierCd", "공급업체코드", "text", "N", "", ""],
        ["status", "상태", "select", "N", "", "DRAFT:임시저장,ACTIVE:활성,INACTIVE:비활성"],
        ["approvalStatus", "승인상태", "select", "N", "", "PENDING:승인대기,APPROVED:승인완료,REJECTED:반려"],
        ["validDateFrom", "유효시작일(부터)", "date", "N", "", ""],
        ["validDateTo", "유효시작일(까지)", "date", "N", "", ""],
    ]
    
    for row in search_conditions:
        ws3.append(row)
    
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Sheet 4: 버튼 정의
    ws4 = wb.create_sheet("04_ButtonDefinitions")
    
    buttons = [
        ["Button ID", "Label", "Style", "Icon", "Action", "Confirm Message"],
        ["btnSearch", "조회", "primary", "bi-search", "search", ""],
        ["btnReset", "초기화", "secondary", "bi-arrow-clockwise", "reset", ""],
        ["btnAdd", "행 추가", "success", "bi-plus-circle", "add", ""],
        ["btnDelete", "행 삭제", "danger", "bi-trash", "delete", "선택한 행을 삭제하시겠습니까?"],
        ["btnSave", "저장", "primary", "bi-save", "save", "저장하시겠습니까?"],
        ["btnApprove", "일괄 승인", "info", "bi-check2-all", "approve", "선택한 항목을 승인하시겠습니까?"],
        ["btnExcelUpload", "Excel 업로드", "warning", "bi-upload", "excelUpload", ""],
        ["btnExcelDownload", "Excel 다운로드", "success", "bi-download", "excelDownload", ""],
        ["btnCopy", "복사", "secondary", "bi-files", "copy", "선택한 행을 복사하시겠습니까?"],
    ]
    
    for row in buttons:
        ws4.append(row)
    
    for cell in ws4[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Sheet 5: API 정의
    ws5 = wb.create_sheet("05_APIDefinitions")
    
    apis = [
        ["API Name", "HTTP Method", "Endpoint", "Description"],
        ["search", "GET", "/api/cost/management/list", "원가 목록 조회"],
        ["save", "POST", "/api/cost/management/save", "원가 저장"],
        ["delete", "DELETE", "/api/cost/management/delete", "원가 삭제"],
        ["approve", "POST", "/api/cost/management/approve", "원가 일괄 승인"],
        ["uploadExcel", "POST", "/api/cost/management/excel/upload", "Excel 업로드"],
        ["downloadExcel", "GET", "/api/cost/management/excel/download", "Excel 다운로드"],
        ["copy", "POST", "/api/cost/management/copy", "원가 복사"],
        ["getFactoryList", "GET", "/api/common/factory/list", "공장 목록 조회"],
        ["getItemList", "GET", "/api/common/item/list", "품목 목록 조회"],
        ["getSupplierList", "GET", "/api/common/supplier/list", "공급업체 목록 조회"],
    ]
    
    for row in apis:
        ws5.append(row)
    
    for cell in ws5[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # 파일 저장
    filename = "/home/roarm_m3/ai-factory-lab/engine/input/CostManagement_ScreenDefinition.xlsx"
    wb.save(filename)
    print(f"✅ Excel PI 생성 완료: {filename}")
    print(f"   - Sheet 1: 기본정보 (15개 항목)")
    print(f"   - Sheet 2: 그리드 컬럼 (31개)")
    print(f"   - Sheet 3: 검색 조건 (12개)")
    print(f"   - Sheet 4: 버튼 정의 (9개)")
    print(f"   - Sheet 5: API 정의 (10개)")

if __name__ == '__main__':
    create_cost_management_pi()
